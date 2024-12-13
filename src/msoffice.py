import pandas as pd
from openpyxl.styles import PatternFill
from docx import Document

from src.address import Address

class MsOffice:
    def __init__(self):
        self.record_per_sheet = 64_000

    def export_to_MS_word(self, address_list: list, file_name: str):
        document = Document()
        table = document.add_table(rows=len(address_list), cols=3)
        table.style = 'Table Grid'
        for i, address in enumerate(address_list):
            row = table.rows[i // 3]  # Divide into 3 columns
            cell = row.cells[i % 3]
            cell.text = address.address
        document.save(file_name)

    def export_to_MS_Excel(self, address_list: list, file_name: str):
        # Prepare data for DataFrame
        data = []
        for address in address_list:
            data.append([
                address.address_old, address.address, address.state, address.district,
                address.block, address.pin, address.phone, "YES" if address.is_reorder else "NO",
                address.name, address.district_from_address, address.state_from_address,
                address.occ_count, address.dist_matches_pin_and_addr, address.state_matches_pin_and_addr,
                address.book_name, address.book_lang, "YES" if address.is_repeat else "NO", address.email, address.faulty
            ])
        
        # Create DataFrame for easier handling
        columns = [
            "ADDRESS ORIGINAL", "ADDRESS UPDATED", "STATE", "DISTRICT", "BLOCK", "PIN", "PHONE", "RE_ORDER",
            "NAME", "DISTRICT_FROM_ADDRESS", "STATE_FROM_ADDRESS", "DISTRICT_MATCH_COUNT",
            "DIST_MATCHES_PIN_AND_ADDR", "STATE_MATCHES_PIN_AND_ADDR", "BOOK NAME", "BOOK LANG", "REPEAT ORDER",
            "EMAIL", "FAULTY"
        ]
        
        df = pd.DataFrame(data, columns=columns)

        # Write to Excel with conditional formatting
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Addresses")
            workbook = writer.book
            worksheet = workbook["Addresses"]

            # Defining styles for the conditional formatting
            style_warn = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            style_alert = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            style_duplicate = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")  # Brown
            style_repeat = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Gray

            # Apply color formatting based on conditions
            for row_idx, address in enumerate(address_list, start=2):  # start=2 to skip header row
                row = worksheet[row_idx]
                
                if address.faulty:
                    # faulty data
                    for col_idx in range(len(row)):
                        row[col_idx].fill = style_warn

                elif address.is_repeat:
                    # Apply "repeat" color
                    for col_idx in range(len(row)):
                        row[col_idx].fill = style_repeat

                elif address.is_reorder:
                    # Apply "duplicate" color
                    for col_idx in range(len(row)):
                        row[col_idx].fill = style_duplicate

                elif address.pin is not None and address.phone is not None:
                    # No specific color; regular address
                    pass

                elif address.phone is not None:
                    # Apply "alert" color
                    for col_idx in range(len(row)):
                        row[col_idx].fill = style_alert

                else:
                    # Apply "warn" color
                    for col_idx in range(len(row)):
                        row[col_idx].fill = style_warn

            # Optional: Set other formatting, like column widths
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

    def import_from_Excel_sheet(self, file_name: str):
        # Read the Excel file into a DataFrame
        df = pd.read_excel(file_name)

        # Convert DataFrame rows into Address objects
        address_list = []
        for _, row in df.iterrows():
            address = Address(
                address_old=row["ADDRESS ORIGINAL"],
                address=row["ADDRESS UPDATED"],
                state=row["STATE"] if pd.notna(row["STATE"]) else None,
                district=row["DISTRICT"] if pd.notna(row["DISTRICT"]) else None,
                block=row["BLOCK"] if pd.notna(row["BLOCK"]) else None,
                pin=row["PIN"] if pd.notna(row["PIN"]) else None,
                phone=row["PHONE"] if pd.notna(row["PHONE"]) else None,
                is_reorder=True if row["RE_ORDER"] == "YES" else False,
                name=row["NAME"],
                district_from_address=row["DISTRICT_FROM_ADDRESS"] if pd.notna(row["DISTRICT_FROM_ADDRESS"]) else None,
                state_from_address=row["STATE_FROM_ADDRESS"] if pd.notna(row["STATE_FROM_ADDRESS"]) else None,
                occ_count=row["DISTRICT_MATCH_COUNT"],
                dist_matches_pin_and_addr=row["DIST_MATCHES_PIN_AND_ADDR"],
                state_matches_pin_and_addr=row["STATE_MATCHES_PIN_AND_ADDR"],
                book_name=row["BOOK NAME"],
                book_lang=row["BOOK LANG"],
                is_repeat=True if row["REPEAT ORDER"] == "YES" else False,
                email=row["EMAIL"],
                faulty=row["FAULTY"]
            )
            address_list.append(address)
        return address_list
